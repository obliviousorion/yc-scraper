"""
scraper/export.py
─────────────────
Export normalised company data to JSON and styled Excel (.xlsx).

The Excel workbook contains two sheets:
  1. **YC Startups** — one row per company with all metadata + links.
  2. **Remote Jobs** — one row per job listing (flattened across companies).

Both sheets are professionally styled with:
  • YC-orange header row with white bold text
  • Alternating row tints for readability
  • Clickable hyperlinks for all URL columns
  • Auto-filters and frozen header panes
"""

import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import log

# ═══════════════════════════════════════════════════════════════════════════════
# JSON
# ═══════════════════════════════════════════════════════════════════════════════

def save_json(data: list[dict], path: str) -> str:
    """Write *data* as pretty-printed JSON and return the path."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    log.info("JSON saved → %s", path)
    return path


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

# ── Column definitions ────────────────────────────────────────────────────────
# Each tuple: (field_key, header_label, column_width)

COMPANY_COLUMNS = [
    ("name",              "Company Name",       28),
    ("short_description", "Short Description",  40),
    ("about",             "About",              60),
    ("batch",             "YC Batch",           12),
    ("status",            "Status",             14),
    ("team_size",         "Team Size",          12),
    ("linkedin_members",  "LinkedIn Members",   18),
    ("remote_jobs_count", "Remote Jobs",        12),
    ("industries",        "Industries",         28),
    ("regions",           "Regions",            28),
    ("tags",              "Tags",               28),
    ("yc_profile_url",    "YC Profile",         40),
    ("yc_jobs_url",       "YC Jobs Page",       40),
    ("website_url",       "Website",            36),
    ("linkedin_url",      "LinkedIn",           36),
    ("twitter_url",       "Twitter / X",        36),
    ("facebook_url",      "Facebook",           36),
    ("github_url",        "GitHub",             36),
    ("crunchbase_url",    "Crunchbase",         36),
    ("youtube_url",       "YouTube",            36),
    ("instagram_url",     "Instagram",          36),
]

COMPANY_URL_KEYS = {
    "yc_profile_url", "yc_jobs_url", "website_url", "linkedin_url",
    "twitter_url", "facebook_url", "github_url", "crunchbase_url",
    "youtube_url", "instagram_url",
}

JOB_COLUMNS = [
    ("company_name",   "Company",         26),
    ("batch",          "YC Batch",        12),
    ("title",          "Job Title",       34),
    ("location",       "Location",        22),
    ("salary",         "Salary",          22),
    ("equity",         "Equity",          16),
    ("experience",     "Experience",      16),
    ("job_url",        "Job Listing",     44),
    ("apply_url",      "Apply Link",      44),
    ("company_yc_url", "Company Profile", 38),
]

JOB_URL_KEYS = {"job_url", "apply_url", "company_yc_url"}

# ── Shared styles ─────────────────────────────────────────────────────────────
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", start_color="F26522")   # YC orange
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT    = Font(name="Arial", size=10)
LINK_FONT    = Font(name="Arial", size=10, color="0563C1", underline="single")
CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)
THIN_BORDER  = Border(
    bottom=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin",  color="E0E0E0"),
)
ALT_FILL = PatternFill("solid", start_color="FFF5EE")  # faint orange tint


def _write_sheet(ws, columns, url_keys, rows, row_height):
    """Write header + data rows onto *ws*, styled consistently."""
    # Header row
    for col_idx, (_, label, width) in enumerate(columns, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, record in enumerate(rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, (key, _, _) in enumerate(columns, start=1):
            value = record.get(key)
            cell  = ws.cell(row=row_idx, column=col_idx)

            if value and key in url_keys:
                cell.value     = value
                cell.hyperlink = value
                cell.font      = LINK_FONT
            else:
                cell.value = value if value is not None else ""
                cell.font  = CELL_FONT

            cell.alignment = CELL_ALIGN
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = row_height

    # Auto-filter
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def save_excel(data: list[dict], path: str) -> str:
    """
    Build and save a styled Excel workbook with two sheets.

    Returns the output path.
    """
    wb = Workbook()

    # Sheet 1 — Companies
    ws = wb.active
    ws.title = "YC Startups"
    _write_sheet(ws, COMPANY_COLUMNS, COMPANY_URL_KEYS, data, row_height=60)

    # Sheet 2 — Remote Jobs (flattened, one row per listing)
    ws_jobs = wb.create_sheet("Remote Jobs")
    job_rows: list[dict] = []
    for startup in data:
        for job in startup.get("remote_jobs", []):
            job_rows.append(
                {
                    "company_name":   startup.get("name", ""),
                    "batch":          startup.get("batch", ""),
                    "title":          job.get("title", ""),
                    "location":       job.get("location", ""),
                    "salary":         job.get("salary", ""),
                    "equity":         job.get("equity", ""),
                    "experience":     job.get("experience", ""),
                    "job_url":        job.get("job_url", ""),
                    "apply_url":      job.get("apply_url", ""),
                    "company_yc_url": startup.get("yc_profile_url", ""),
                }
            )
    _write_sheet(ws_jobs, JOB_COLUMNS, JOB_URL_KEYS, job_rows, row_height=32)

    wb.save(path)
    log.info(
        "Excel saved → %s  (%d companies, %d remote jobs)",
        path,
        len(data),
        len(job_rows),
    )
    return path
